#Report Manager
from src.service import *
from src.kpi_service import *
from datetime import datetime
from openpyxl import Workbook


class ReportManager:
    def __init__(self,config,data_source):
        self.config=config
        self.data_source = data_source
        self.report_folder= config["report_folder"]
        self.database_file = config["database_file"]

    def generate_csv_report(self):
        loginfo_print("Generating CSV report...")
    
    def generate_summary(self):
        loginfo_print("Generating KPI summary...")

    def create_report_name(self, file_ext):
        return f"{self.report_folder}/KPI_report_{datetime.now().strftime('%Y_%m_%d_%H_%M')}.{file_ext}"
    
    def generate_excel_report(self):

        try:
            total_records = self.data_source.get_total_records() #get_total_records(self.database_file)
            running_records = self.data_source.get_running_machines() #get_running_machines(self.database_file)
            total_alarms = self.data_source.get_total_alarms() #get_total_alarms(self.database_file)
            critical_alarms = self.data_source.get_critical_alarms() #get_critical_alarms(self.database_file)
            alarm_summary = self.data_source.get_alarm_summary() #get_alarm_summary(self.database_file)
            availability = (running_records / total_records)*100 if total_records > 0 else 0

            factory_inventory = self.data_source.get_factory_inventory() #get_factory_inventory(self.database_file)
        except Exception as e:
            logerror_print("Error occurred while fetching KPI data: " + str(e))
            total_records = 0
            total_alarms = 0
            critical_alarms = 0
            running_records = 0
            alarm_summary = 0
            availability = 0

        

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Summary"

        sheet["A1"] = "Industrial Monitor KPI Report"
        sheet["A3"] = "Generated at"
        sheet["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        sheet["A5"] = "Total Machines"
        sheet["B5"] = total_records

        sheet["A6"] = "Total Alarms"
        sheet["B6"] = total_alarms

        sheet["A7"] = "Critical Alarms"
        sheet["B7"] = critical_alarms

        sheet["A8"] = "Availability (%)"
        sheet["B8"] = f"{availability:.2f}%"

        workbook.create_sheet(title="Alarm Summary")
        sheet2 = workbook["Alarm Summary"]

        sheet2["A3"] = "Alarm Type"
        sheet2["B3"] = "Count"
    
        try:
            row = 4

            for alarm_type, count in alarm_summary:
                sheet2[f"A{row}"] = alarm_type
                sheet2[f"B{row}"] = count
                row += 1
        except Exception as error:
            logerror_print("Error occurred while generating alarm: " + str(error))

        filename = self.create_report_name("xlsx")

        workbook.create_sheet(title="Factory Inventory")
        sheet3 = workbook["Factory Inventory"]

        sheet3["A3"]= "Machine ID"
        sheet3["B3"]= "Machine Type"
        sheet3["C3"]= "Status"

        try:
            row1 = 4

            for machine_id, machine_type, status in factory_inventory:
                sheet3[f"A{row1}"] = machine_id
                sheet3[f"B{row1}"] = machine_type
                sheet3[f"C{row1}"] = status
                row1 +=1
        except Exception as error:
            logerror_print("Error occurred while generating factory inventory: " + str(error))


        workbook.save(filename)
        loginfo_print("Excel report generated: "+ filename)




